import customtkinter as ctk
import openpyxl
from pylatex import Document, Math, Section


class FileExporter:
    def __init__(self, translator) -> None:
        self.translator = translator

    def export_to_excel(self, result):
        wb = openpyxl.Workbook()
        ws = wb.active
        if self.translator.language == "pl":
            label = "Wynik równania:"
        else:
            label = "Solution:"
        ws["A1"] = label
        ws["A2"] = result

        file_path = self.ask_save_file(
            defaultextension=".xlsx",
            filetypes=[("Pliki Excela", "*.xlsx")],
            title="Zapisz plik Excela",
        )
        if file_path:
            wb.save(file_path)

    def export_to_latex(self, result):
        doc = Document()
        if self.translator.language == "pl":
            label = "Wynik równania:"
        else:
            label = "Solution:"
        with doc.create(Section(label)):
            with doc.create(Math()):
                doc.append(result)

        file_path = self.ask_save_file(
            defaultextension=".tex",
            filetypes=[("Pliki LaTeX", "*.tex")],
            title="Zapisz plik LaTeX",
        )
        if file_path:
            doc.generate_tex(file_path)

    def ask_save_file(self, defaultextension, filetypes, title):
        return ctk.filedialog.asksaveasfilename(
            defaultextension=defaultextension,
            filetypes=filetypes,
            title=title,
        )
